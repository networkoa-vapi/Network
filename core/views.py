from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.urls import reverse
from django.core.exceptions import ValidationError
from .models import Quotation, ServiceTicket, CustomerProfile, Company, Product, Division, ProductCategory, ProductSubCategory
from django.http import HttpResponseForbidden

def _from_email_for(request):
    """Documents emailed from the CRM go out as the logged-in user's own email (each user
    configures theirs under Master › Users) - falls back to the site default if unset."""
    email = getattr(request.user, 'email', '') or ''
    return email.strip() or None

def _pdf_link_callback(uri, rel):
    """Resolves {% static %}/{% media %} URLs used inside a PDF template (e.g. a
    letterhead image) into real filesystem paths, since xhtml2pdf can't fetch them itself."""
    import os
    from django.conf import settings
    from django.contrib.staticfiles import finders

    if uri.startswith(settings.STATIC_URL):
        path = finders.find(uri[len(settings.STATIC_URL):])
        if path:
            return path
    if uri.startswith(settings.MEDIA_URL):
        path = os.path.join(settings.MEDIA_ROOT, uri[len(settings.MEDIA_URL):])
        if os.path.isfile(path):
            return path
    return uri

def is_customer(user):
    return user.is_authenticated and user.role and user.role.name == 'Customer'

def is_engineer(user):
    return user.is_authenticated and user.role and user.role.name == 'Engineer'

def portal_login(request):
    if request.user.is_authenticated:
        if request.user.role and request.user.role.name == 'Customer':
            return redirect('dashboard')
        elif request.user.role and request.user.role.name == 'Engineer':
            return redirect('engineer_dashboard')
        else:
            return redirect('/admin/')
            
    if request.method == 'POST':
        u = request.POST.get('username')
        p = request.POST.get('password')
        user = authenticate(request, username=u, password=p)
        if user is not None:
            if user.role and user.role.name == 'Customer':
                login(request, user)
                return redirect('dashboard')
            elif user.role and user.role.name == 'Engineer':
                login(request, user)
                return redirect('engineer_dashboard')
            else:
                login(request, user)
                return redirect('/admin/')
        else:
            messages.error(request, 'Invalid credentials.')
            
    return render(request, 'portal/login.html')

def portal_logout(request):
    logout(request)
    return redirect('portal_login')

@login_required
def dashboard(request):
    if not is_customer(request.user): return redirect('/admin/')
    
    try:
        profile = request.user.customerprofile
    except:
        return HttpResponseForbidden("Customer profile not found.")
        
    tickets_count = ServiceTicket.objects.filter(customer=profile).exclude(status='closed').count()
    quotes_count = Quotation.objects.filter(customer=profile, status='sent').count()
    
    context = {
        'tickets_count': tickets_count,
        'quotes_count': quotes_count,
        'profile': profile
    }
    return render(request, 'portal/dashboard.html', context)

@login_required
def quotations_list(request):
    if not is_customer(request.user): return redirect('/admin/')
    profile = request.user.customerprofile
    # Only show quotes that are sent or accepted/rejected
    quotes = Quotation.objects.filter(customer=profile, status__in=['sent', 'accepted', 'rejected']).order_by('-created_at')
    
    return render(request, 'portal/quotations.html', {'quotes': quotes})

@login_required
def accept_quotation(request, quote_id):
    if not is_customer(request.user): return redirect('/admin/')
    if request.method == 'POST':
        profile = request.user.customerprofile
        quote = get_object_or_404(Quotation, id=quote_id, customer=profile, status='sent')
        quote.status = 'accepted'
        quote.save()
        first_item = quote.items.first()
        product_name = first_item.product.name if first_item and first_item.product else "your requested products"
        messages.success(request, f'Quotation for {product_name} has been accepted.')
    return redirect('quotations_list')

@login_required
def tickets_list(request):
    if not is_customer(request.user): return redirect('/admin/')
    profile = request.user.customerprofile
    tickets = ServiceTicket.objects.filter(customer=profile).order_by('-created_at')
    return render(request, 'portal/tickets.html', {'tickets': tickets})

@login_required
def create_ticket(request):
    if not is_customer(request.user): return redirect('/admin/')
    profile = request.user.customerprofile
    
    if request.method == 'POST':
        title = request.POST.get('title')
        desc = request.POST.get('description')
        priority = request.POST.get('priority')
        product_id = request.POST.get('product')
        division_id = request.POST.get('division')
        category_id = request.POST.get('category')
        subcategory_id = request.POST.get('subcategory')
        
        product = Product.objects.get(id=product_id) if product_id else None
        
        ticket = ServiceTicket.objects.create(
            company=profile.company,
            customer=profile,
            product=product,
            issue_title=title,
            description=desc,
            priority=priority
        )
        
        if division_id:
            try:
                ticket.division = Division.objects.get(id=division_id)
            except Division.DoesNotExist:
                pass
        if category_id:
            try:
                ticket.category = ProductCategory.objects.get(id=category_id)
            except ProductCategory.DoesNotExist:
                pass
        if subcategory_id:
            try:
                ticket.subcategory = ProductSubCategory.objects.get(id=subcategory_id)
            except ProductSubCategory.DoesNotExist:
                pass
        ticket.save()
        
        messages.success(request, 'Your service ticket has been submitted successfully.')
        return redirect('tickets_list')
        
    divisions = Division.objects.filter(company=profile.company)
    return render(request, 'portal/ticket_create.html', {'divisions': divisions})

# ==========================================
# ENGINEER PORTAL VIEWS
# ==========================================

@login_required
def engineer_dashboard(request):
    if not is_engineer(request.user): return redirect('/admin/')
    
    try:
        employee = request.user.employee
    except:
        return HttpResponseForbidden("Employee profile not found.")
        
    assigned_tickets = ServiceTicket.objects.filter(assigned_engineers=employee).exclude(status__in=['resolved', 'closed']).count()
    resolved_tickets = ServiceTicket.objects.filter(assigned_engineers=employee, status='resolved').count()
    
    context = {
        'employee': employee,
        'assigned_tickets': assigned_tickets,
        'resolved_tickets': resolved_tickets,
    }
    return render(request, 'engineer/dashboard.html', context)

@login_required
def engineer_tickets(request):
    if not is_engineer(request.user): return redirect('/admin/')
    employee = request.user.employee
    
    # Show active tickets first, then resolved/closed
    tickets = ServiceTicket.objects.filter(assigned_engineers=employee).order_by('status', '-created_at')
    return render(request, 'engineer/tickets.html', {'tickets': tickets})

@login_required
def engineer_ticket_update(request, ticket_id):
    if not is_engineer(request.user): return redirect('/admin/')
    employee = request.user.employee
    
    ticket = get_object_or_404(ServiceTicket, id=ticket_id, assigned_engineers=employee)

    if request.method == 'POST':
        new_status = request.POST.get('status')
        notes = request.POST.get('resolution_notes')

        if new_status:
            ticket.status = new_status
        if notes:
            ticket.resolution_notes = notes

        try:
            ticket.full_clean()
            ticket.save()
            messages.success(request, f'Ticket #{ticket.id} has been updated successfully.')
        except ValidationError as e:
            messages.error(request, ' '.join(e.messages))
        return redirect('engineer_tickets')
        
    return render(request, 'engineer/ticket_update.html', {'ticket': ticket})

# ==========================================
# PDF GENERATION VIEWS
# ==========================================
from django.template.loader import get_template
from django.http import HttpResponse
from django.utils import timezone
from xhtml2pdf import pisa
from io import BytesIO
from .models import Quotation, Invoice

@login_required
def generate_pdf_quotation(request, quote_id):
    # Admins or the specific customer can download
    quote = get_object_or_404(Quotation, id=quote_id)
    
    # Check permission
    if not (request.user.is_superuser or (is_customer(request.user) and quote.customer == request.user.customerprofile)):
        return HttpResponseForbidden("You do not have permission to download this quotation.")
        
    template = get_template('pdf/quotation_pdf.html')
    
    # Calculate discount amount
    subtotal = quote.get_subtotal()
    discount_amount = subtotal * (quote.admin_discount_percent / 100)
    
    context = {
        'quotation': quote,
        'discount_amount': discount_amount
    }
    
    html = template.render(context)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)
    
    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        filename = f"Quotation_{quote.quotation_number}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    return HttpResponse("Error generating PDF", status=400)

@login_required
def generate_pdf_invoice(request, invoice_id):
    invoice = get_object_or_404(Invoice, id=invoice_id)
    template = get_template('pdf/invoice_pdf.html')
    context = {'invoice': invoice}
    
    html = template.render(context)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)
    
    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        filename_num = invoice.invoice_number if invoice.invoice_number else invoice.id
        response['Content-Disposition'] = f'attachment; filename="Invoice_{filename_num}.pdf"'
        return response
    return HttpResponse("Error generating PDF", status=400)

@login_required
def generate_pdf_stock_report(request):
    if not request.user.is_staff:
        return HttpResponseForbidden("You do not have permission to view this report.")

    from datetime import datetime, timedelta
    from django.db.models import Sum
    from .models import Godown, StockItem

    godown_id = request.GET.get('godown')
    ids_param = request.GET.get('ids')
    category_id = request.GET.get('category')
    subcategory_id = request.GET.get('subcategory')
    unit = request.GET.get('unit')
    low_stock_only = request.GET.get('low_stock_only') == '1'

    date_from = date_to = None
    date_from_raw = request.GET.get('date_from', '').strip()
    date_to_raw = request.GET.get('date_to', '').strip()
    if date_from_raw:
        try:
            date_from = datetime.strptime(date_from_raw, '%Y-%m-%d').date()
        except ValueError:
            date_from = None
    if date_to_raw:
        try:
            date_to = datetime.strptime(date_to_raw, '%Y-%m-%d').date()
        except ValueError:
            date_to = None

    items = StockItem.objects.filter(is_active=True).select_related('category', 'subcategory')
    if ids_param:
        items = items.filter(pk__in=[i for i in ids_param.split(',') if i.isdigit()])
    if category_id:
        items = items.filter(category_id=category_id)
    if subcategory_id:
        items = items.filter(subcategory_id=subcategory_id)
    if unit:
        items = items.filter(unit=unit)

    godowns = Godown.objects.filter(is_active=True).order_by('godown_type', 'name')
    if godown_id:
        godowns = godowns.filter(pk=godown_id)

    def movement(item, godown):
        """Opening / Purchased / Issued / Closing for this item over [date_from, date_to] -
        blank dates mean "since the beginning" / "as of right now", so the existing
        no-filters report still just shows current totals with Opening at 0."""
        opening = item.get_stock_balance(godown=godown, as_of=date_from - timedelta(days=1)) if date_from else 0
        txns = item.transactions.all()
        if godown is not None:
            txns = txns.filter(godown=godown)
        if date_from:
            txns = txns.filter(transaction_date__gte=date_from)
        if date_to:
            txns = txns.filter(transaction_date__lte=date_to)
        purchased = txns.filter(transaction_type='receipt').aggregate(total=Sum('quantity'))['total'] or 0
        issued = txns.filter(transaction_type='issue').aggregate(total=Sum('quantity'))['total'] or 0
        closing = item.get_stock_balance(godown=godown, as_of=date_to) if date_to else item.get_stock_balance(godown=godown)
        return opening, purchased, issued, closing

    def piece_breakdown(item, godown):
        """Current in-stock pieces for a piece-tracked item (e.g. Copper Pipe cut lengths),
        so the report can show the total alongside how it's actually broken up."""
        if not item.is_piece_tracked:
            return None
        qs = item.pieces.filter(status='in_stock')
        if godown is not None:
            qs = qs.filter(godown=godown)
        return list(qs.order_by('-quantity'))

    summary_rows = []
    for item in items:
        opening, purchased, issued, closing = movement(item, None)
        low_stock = closing <= item.reorder_level
        if low_stock_only and not low_stock:
            continue
        summary_rows.append({
            'item': item, 'opening': opening, 'purchased': purchased, 'issued': issued,
            'total': closing, 'low_stock': low_stock, 'pieces': piece_breakdown(item, None),
        })

    godown_sections = []
    for godown in godowns:
        rows = []
        for item in items:
            opening, purchased, issued, closing = movement(item, godown)
            if not (opening or purchased or issued or closing):
                continue
            low_stock = closing <= item.reorder_level
            if low_stock_only and not low_stock:
                continue
            rows.append({
                'item': item, 'opening': opening, 'purchased': purchased, 'issued': issued,
                'balance': closing, 'low_stock': low_stock, 'pieces': piece_breakdown(item, godown),
            })
        godown_sections.append({'godown': godown, 'rows': rows})

    template = get_template('pdf/stock_report_pdf.html')
    context = {
        'company': Company.objects.order_by('pk').first(),
        'generated_at': timezone.now(),
        'summary_rows': summary_rows,
        'godown_sections': godown_sections,
        'filtered_godown': godowns.first() if godown_id else None,
        'date_from': date_from,
        'date_to': date_to,
    }

    html = template.render(context)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)

    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="Stock_Report.pdf"'
        return response
    return HttpResponse("Error generating PDF", status=400)

@login_required
def stock_report_filters_view(request):
    if not request.user.is_staff:
        return HttpResponseForbidden("You do not have permission to view this report.")

    from .models import Godown, StockCategory, StockSubCategory, StockItem

    context = {
        'godowns': Godown.objects.filter(is_active=True).order_by('godown_type', 'name'),
        'categories': StockCategory.objects.filter(is_active=True).order_by('name'),
        'subcategories': StockSubCategory.objects.filter(is_active=True).select_related('category').order_by('category__name', 'name'),
        'units': StockItem.UNIT_CHOICES,
    }
    return render(request, 'admin/store/stock_report_filters.html', context)

# ==========================================
# PROJECT DEPARTMENT - PURCHASE REQUISITION
# ==========================================

@login_required
def stock_item_balance_api(request, item_id):
    from django.http import JsonResponse
    from .models import StockItem
    if not request.user.is_staff:
        return JsonResponse({'error': 'forbidden'}, status=403)
    item = get_object_or_404(StockItem, pk=item_id)
    return JsonResponse({
        'balance': str(item.get_stock_balance()),
        'reorder_level': str(item.reorder_level),
        'unit': item.get_unit_display(),
        'unit_code': item.unit,
        'is_serialized': item.is_serialized,
        'is_returnable': item.is_returnable,
        'is_refillable': item.is_refillable,
        'is_piece_tracked': item.is_piece_tracked,
    })

@login_required
def stock_item_serials_api(request, item_id):
    from django.http import JsonResponse
    from .models import StockItemSerial
    if not request.user.is_staff:
        return JsonResponse({'error': 'forbidden'}, status=403)
    statuses = [s for s in request.GET.get('statuses', 'in_stock').split(',') if s]
    serials = StockItemSerial.objects.filter(stock_item_id=item_id, status__in=statuses).order_by('serial_number')
    return JsonResponse([
        {'id': s.id, 'serial_number': s.serial_number, 'status': s.status}
        for s in serials
    ], safe=False)

@login_required
def stock_item_pieces_api(request, item_id):
    from django.http import JsonResponse
    from .models import StockItemPiece
    if not request.user.is_staff:
        return JsonResponse({'error': 'forbidden'}, status=403)
    statuses = [s for s in request.GET.get('statuses', 'in_stock').split(',') if s]
    pieces = StockItemPiece.objects.filter(stock_item_id=item_id, status__in=statuses).order_by('-quantity')
    return JsonResponse([
        {'id': p.id, 'label': p.label, 'quantity': str(p.quantity), 'status': p.status}
        for p in pieces
    ], safe=False)

@login_required
def generate_pdf_purchase_requisition(request, pr_id):
    from .models import PurchaseRequisition
    if not request.user.is_staff:
        return HttpResponseForbidden("You do not have permission to view this requisition.")

    pr = get_object_or_404(PurchaseRequisition, id=pr_id)
    template = get_template('pdf/purchase_requisition_pdf.html')
    context = {
        'company': pr.company,
        'pr': pr,
        'generated_at': timezone.now(),
    }

    html = template.render(context)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)

    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="Purchase_Requisition_{pr.pr_number}.pdf"'
        return response
    return HttpResponse("Error generating PDF", status=400)

@login_required
def email_purchase_requisition(request, pr_id):
    from django.core.mail import EmailMessage
    from .models import PurchaseRequisition

    if not request.user.is_staff:
        return HttpResponseForbidden("You do not have permission to email this requisition.")

    pr = get_object_or_404(PurchaseRequisition, id=pr_id)

    if request.method == 'POST':
        to_email = request.POST.get('email', '').strip()
        note = request.POST.get('note', '').strip()
        if not to_email:
            messages.error(request, 'Please enter an email address.')
            return redirect('email_purchase_requisition', pr_id=pr.id)

        template = get_template('pdf/purchase_requisition_pdf.html')
        context = {'company': pr.company, 'pr': pr, 'generated_at': timezone.now()}
        html = template.render(context)
        result = BytesIO()
        pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)

        if pdf.err:
            messages.error(request, 'Could not generate the PDF for this requisition.')
            return redirect('email_purchase_requisition', pr_id=pr.id)

        subject = f'Purchase Requisition {pr.pr_number} - {pr.project.project_number}'
        body = note or f'Please find attached Purchase Requisition {pr.pr_number} for project {pr.project.project_number} ({pr.project.customer.business_name}).'
        email = EmailMessage(subject, body, from_email=_from_email_for(request), to=[to_email])
        email.attach(f'Purchase_Requisition_{pr.pr_number}.pdf', result.getvalue(), 'application/pdf')
        try:
            email.send(fail_silently=False)
            messages.success(request, f'Requisition emailed to {to_email}.')
        except Exception as e:
            messages.error(request, f'Could not send email - check email server settings. ({e})')
        return redirect('admin:project_purchaserequisition_change', pr.id)

    return render(request, 'admin/project/email_purchase_requisition.html', {'pr': pr})

# ==========================================
# PURCHASE DEPARTMENT - SUPPLIER PURCHASE ORDER
# ==========================================

@login_required
def generate_pdf_supplier_po(request, po_id):
    from .models import SupplierPurchaseOrder
    if not request.user.is_staff:
        return HttpResponseForbidden("You do not have permission to view this purchase order.")

    po = get_object_or_404(SupplierPurchaseOrder, id=po_id)
    template = get_template('pdf/supplier_po_pdf.html')
    context = {'company': po.company, 'po': po, 'generated_at': timezone.now()}

    html = template.render(context)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)

    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="Purchase_Order_{po.po_number}.pdf"'
        return response
    return HttpResponse("Error generating PDF", status=400)

@login_required
def email_supplier_po(request, po_id):
    from django.core.mail import EmailMessage
    from .models import SupplierPurchaseOrder

    if not request.user.is_staff:
        return HttpResponseForbidden("You do not have permission to email this purchase order.")

    po = get_object_or_404(SupplierPurchaseOrder, id=po_id)

    if request.method == 'POST':
        to_email = request.POST.get('email', '').strip() or po.supplier.email
        note = request.POST.get('note', '').strip()
        if not to_email:
            messages.error(request, 'Please enter an email address (this supplier has none on file).')
            return redirect('email_supplier_po', po_id=po.id)

        template = get_template('pdf/supplier_po_pdf.html')
        context = {'company': po.company, 'po': po, 'generated_at': timezone.now()}
        html = template.render(context)
        result = BytesIO()
        pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)

        if pdf.err:
            messages.error(request, 'Could not generate the PDF for this purchase order.')
            return redirect('email_supplier_po', po_id=po.id)

        subject = f'Purchase Order {po.po_number} from {po.company.name}'
        body = note or f'Dear {po.supplier.name},\n\nPlease find attached Purchase Order {po.po_number}.\n\nRegards,\n{po.company.name}'
        email = EmailMessage(subject, body, from_email=_from_email_for(request), to=[to_email])
        email.attach(f'Purchase_Order_{po.po_number}.pdf', result.getvalue(), 'application/pdf')
        try:
            email.send(fail_silently=False)
            if po.status == 'draft':
                po.status = 'sent'
                po.save()
            messages.success(request, f'Purchase Order emailed to {to_email}.')
        except Exception as e:
            messages.error(request, f'Could not send email - check email server settings. ({e})')
        return redirect('admin:purchase_supplierpurchaseorder_change', po.id)

    return render(request, 'admin/purchase/email_supplier_po.html', {'po': po})

@login_required
def whatsapp_share_supplier_po(request, po_id):
    from urllib.parse import quote
    from .models import SupplierPurchaseOrder

    if not request.user.is_staff:
        return HttpResponseForbidden("You do not have permission to share this purchase order.")

    po = get_object_or_404(SupplierPurchaseOrder, id=po_id)
    digits = ''.join(ch for ch in po.supplier.whatsapp_number if ch.isdigit())
    if not digits:
        messages.error(request, f"{po.supplier.name} has no WhatsApp number on file - add one in the Supplier record first.")
        return redirect('admin:purchase_supplierpurchaseorder_change', po.id)

    pdf_url = request.build_absolute_uri(reverse('generate_pdf_supplier_po', args=[po.id]))
    message = f"Purchase Order {po.po_number} from {po.company.name}. View/download: {pdf_url}"

    if po.status == 'draft':
        po.status = 'sent'
        po.save()

    return redirect(f"https://wa.me/{digits}?text={quote(message)}")

# ==========================================
# SERVICE HUB - EQUIPMENT REPORT
# ==========================================

@login_required
def generate_pdf_equipment_report(request):
    from .models import Equipment
    if not request.user.is_staff:
        return HttpResponseForbidden("You do not have permission to view this report.")

    ids_param = request.GET.get('ids')
    equipment = Equipment.objects.select_related('customer', 'category', 'subcategory').order_by('customer__business_name')
    if ids_param:
        equipment = equipment.filter(pk__in=[i for i in ids_param.split(',') if i.isdigit()])

    template = get_template('pdf/equipment_report_pdf.html')
    context = {
        'company': Company.objects.order_by('pk').first(),
        'generated_at': timezone.now(),
        'equipment_list': equipment,
        'is_single': equipment.count() == 1,
    }

    html = template.render(context)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)

    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="Equipment_Report.pdf"'
        return response
    return HttpResponse("Error generating PDF", status=400)

# ==========================================
# HR & ADMIN - LETTER GENERATION
# ==========================================

@login_required
def generate_pdf_hr_letter(request, employee_id, letter_type):
    from .models import Employee
    from django.http import Http404
    if not request.user.is_staff:
        return HttpResponseForbidden("You do not have permission to generate this letter.")

    letter_titles = {
        'experience': 'Experience Letter',
        'termination': 'Termination Letter',
    }
    if letter_type not in letter_titles:
        raise Http404("Unknown letter type.")

    employee = get_object_or_404(Employee, id=employee_id)
    template = get_template('pdf/hr_letter_pdf.html')
    context = {
        'company': employee.company,
        'employee': employee,
        'letter_type': letter_type,
        'letter_title': letter_titles[letter_type],
        'generated_at': timezone.now(),
    }

    html = template.render(context)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)

    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="{letter_titles[letter_type].replace(" ", "_")}_{employee.employee_code}.pdf"'
        return response
    return HttpResponse("Error generating PDF", status=400)

# ==========================================
# HR & ADMIN - OFFER LETTER (matches the company's own letterhead pad)
# ==========================================

def _render_offer_letter_pdf(offer):
    template = get_template('pdf/offer_letter_pdf.html')
    context = {'company': offer.company, 'offer': offer, 'generated_at': timezone.now()}
    html = template.render(context)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result, link_callback=_pdf_link_callback)
    return None if pdf.err else result.getvalue()

@login_required
def generate_pdf_offer_letter(request, offer_id):
    from .models import OfferLetter
    if not request.user.is_staff:
        return HttpResponseForbidden("You do not have permission to view this letter.")

    offer = get_object_or_404(OfferLetter, id=offer_id)
    pdf_bytes = _render_offer_letter_pdf(offer)
    if pdf_bytes is None:
        return HttpResponse("Error generating PDF", status=400)
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="Offer_Letter_{offer.candidate_name.replace(" ", "_")}.pdf"'
    return response

@login_required
def email_offer_letter(request, offer_id):
    from django.core.mail import EmailMessage
    from .models import OfferLetter

    if not request.user.is_staff:
        return HttpResponseForbidden("You do not have permission to email this letter.")

    offer = get_object_or_404(OfferLetter, id=offer_id)

    if request.method == 'POST':
        to_email = request.POST.get('email', '').strip()
        note = request.POST.get('note', '').strip()
        if not to_email:
            messages.error(request, 'Please enter an email address.')
            return redirect('email_offer_letter', offer_id=offer.id)

        pdf_bytes = _render_offer_letter_pdf(offer)
        if pdf_bytes is None:
            messages.error(request, 'Could not generate the PDF for this letter.')
            return redirect('email_offer_letter', offer_id=offer.id)

        subject = f'Offer Letter {offer.letter_number} - {offer.company.name}'
        body = note or f'Dear {offer.candidate_title} {offer.candidate_first_name},\n\nPlease find attached your Offer Letter ({offer.letter_number}) for the position of {offer.designation}.\n\nRegards,\n{offer.company.name}'
        email = EmailMessage(subject, body, from_email=_from_email_for(request), to=[to_email])
        email.attach(f'Offer_Letter_{offer.candidate_name.replace(" ", "_")}.pdf', pdf_bytes, 'application/pdf')
        try:
            email.send(fail_silently=False)
            messages.success(request, f'Offer Letter emailed to {to_email}.')
        except Exception as e:
            messages.error(request, f'Could not send email - check email server settings. ({e})')
        return redirect('admin:hr_hrofferletter_change', offer.id)

    return render(request, 'admin/hr/email_offer_letter.html', {'offer': offer, 'default_email': offer.candidate_email})

@login_required
def whatsapp_share_offer_letter(request, offer_id):
    import re
    from urllib.parse import quote
    from .models import OfferLetter

    if not request.user.is_staff:
        return HttpResponseForbidden("You do not have permission to share this letter.")

    offer = get_object_or_404(OfferLetter, id=offer_id)
    pdf_url = request.build_absolute_uri(reverse('generate_pdf_offer_letter', args=[offer.id]))
    message = f"Offer Letter {offer.letter_number} for {offer.candidate_name} ({offer.designation}) - {offer.company.name}. View/download: {pdf_url}"
    digits = re.sub(r'\D', '', offer.candidate_phone or '')
    target = f"https://wa.me/{digits}" if digits else "https://wa.me/"
    return redirect(f"{target}?text={quote(message)}")

# ==========================================
# UNIVERSAL REPORTS CENTER
# One report picker for every department - flexible date range, export to
# PDF / Excel / CSV, and share by Email or WhatsApp.
# ==========================================

DURATION_PRESETS = (
    ('full', 'Full / All Time'),
    ('today', 'Today'),
    ('yesterday', 'Yesterday'),
    ('this_week', 'This Week'),
    ('this_month', 'This Month'),
    ('this_quarter', 'This Quarter'),
    ('this_year', 'This Year'),
    ('last_30_days', 'Last 30 Days'),
    ('last_90_days', 'Last 90 Days'),
    ('custom', 'Custom Range'),
)

FORMAT_CHOICES = (
    ('pdf', 'PDF'),
    ('xlsx', 'Excel (XLSX)'),
    ('csv', 'CSV'),
)


def _resolve_date_range(preset, date_from, date_to):
    """Returns (start_date, end_date) - both None for 'full' (no filtering)."""
    from datetime import datetime, timedelta
    today = timezone.now().date()
    if preset == 'today':
        return today, today
    if preset == 'yesterday':
        y = today - timedelta(days=1)
        return y, y
    if preset == 'this_week':
        return today - timedelta(days=today.weekday()), today
    if preset == 'this_month':
        return today.replace(day=1), today
    if preset == 'this_quarter':
        q_start_month = ((today.month - 1) // 3) * 3 + 1
        return today.replace(month=q_start_month, day=1), today
    if preset == 'this_year':
        return today.replace(month=1, day=1), today
    if preset == 'last_30_days':
        return today - timedelta(days=30), today
    if preset == 'last_90_days':
        return today - timedelta(days=90), today
    if preset == 'custom':
        df = dt = None
        if date_from:
            try:
                df = datetime.strptime(date_from, '%Y-%m-%d').date()
            except ValueError:
                pass
        if date_to:
            try:
                dt = datetime.strptime(date_to, '%Y-%m-%d').date()
            except ValueError:
                pass
        return df, dt
    return None, None


def _get_report_data(params):
    """report key + duration params -> (report_key, report_def, headers, rows, date_range_label) or all-None if invalid."""
    from datetime import timedelta
    from .reports import REPORTS

    report_key = params.get('report')
    report_def = REPORTS.get(report_key)
    if not report_def:
        return None, None, None, None, None

    qs = report_def['model'].objects.all()
    date_field = report_def.get('date_field')
    preset = params.get('preset', 'full')
    date_range_label = 'Full / All Time'

    if date_field and preset != 'full':
        start, end = _resolve_date_range(preset, params.get('date_from'), params.get('date_to'))
        if start:
            qs = qs.filter(**{f'{date_field}__gte': start})
        if end:
            qs = qs.filter(**{f'{date_field}__lt': end + timedelta(days=1)})
        if start or end:
            date_range_label = f"{start or '...'} to {end or '...'}"

    headers = [c[0] for c in report_def['columns']]
    rows = [[col[1](obj) for col in report_def['columns']] for obj in qs]
    return report_key, report_def, headers, rows, date_range_label


def _excel_safe(value):
    from decimal import Decimal
    if isinstance(value, Decimal):
        return float(value)
    if value is None:
        return ''
    return value


def _build_xlsx_bytes(sheet_title, headers, rows):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_title or 'Report')[:31]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append([_excel_safe(v) for v in row])
    for i, header in enumerate(headers, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(12, len(str(header)) + 2)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_csv_text(headers, rows):
    import csv
    from io import StringIO

    buf = StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(['' if v is None else v for v in row])
    return buf.getvalue()


def _build_report_pdf_bytes(report_def, headers, rows, date_range_label):
    template = get_template('pdf/generic_report_pdf.html')
    context = {
        'company': Company.objects.order_by('pk').first(),
        'generated_at': timezone.now(),
        'report_label': report_def['label'],
        'department': report_def['department'],
        'date_range_label': date_range_label,
        'headers': headers,
        'rows': rows,
    }
    html = template.render(context)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)
    return None if pdf.err else result.getvalue()


@login_required
def reports_center(request):
    if not request.user.is_staff:
        return HttpResponseForbidden("You do not have permission to view reports.")
    from .reports import REPORTS, REPORT_DEPARTMENT_ORDER

    departments = []
    for dept in REPORT_DEPARTMENT_ORDER:
        options = [(key, r['label']) for key, r in REPORTS.items() if r['department'] == dept]
        if options:
            departments.append((dept, options))

    context = {
        'departments': departments,
        'duration_presets': DURATION_PRESETS,
        'format_choices': FORMAT_CHOICES,
    }
    return render(request, 'admin/reports_center.html', context)


def _report_filename(report_def, fmt):
    return f"{report_def['label'].replace(' ', '_')}.{fmt}"


@login_required
def reports_download(request):
    if not request.user.is_staff:
        return HttpResponseForbidden("You do not have permission to view reports.")

    report_key, report_def, headers, rows, date_range_label = _get_report_data(request.GET)
    if not report_def:
        return HttpResponse("Unknown report selected.", status=400)
    fmt = request.GET.get('format', 'pdf')

    if fmt == 'xlsx':
        content = _build_xlsx_bytes(report_def['label'], headers, rows)
        response = HttpResponse(content, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{_report_filename(report_def, "xlsx")}"'
        return response

    if fmt == 'csv':
        content = _build_csv_text(headers, rows)
        response = HttpResponse(content, content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{_report_filename(report_def, "csv")}"'
        return response

    pdf_bytes = _build_report_pdf_bytes(report_def, headers, rows, date_range_label)
    if pdf_bytes is None:
        return HttpResponse("Error generating PDF", status=400)
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="{_report_filename(report_def, "pdf")}"'
    return response


@login_required
def reports_email(request):
    if not request.user.is_staff:
        return HttpResponseForbidden("You do not have permission to email reports.")
    if request.method != 'POST':
        return HttpResponseForbidden("Invalid request method.")

    from django.core.mail import EmailMessage

    report_key, report_def, headers, rows, date_range_label = _get_report_data(request.POST)
    if not report_def:
        messages.error(request, 'Please choose a report first.')
        return redirect('reports_center')

    to_email = request.POST.get('email', '').strip()
    if not to_email:
        messages.error(request, 'Please enter a recipient email address.')
        return redirect('reports_center')

    fmt = request.POST.get('format', 'pdf')
    if fmt == 'xlsx':
        content = _build_xlsx_bytes(report_def['label'], headers, rows)
        mime, ext = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'xlsx'
    elif fmt == 'csv':
        content = _build_csv_text(headers, rows).encode('utf-8')
        mime, ext = 'text/csv', 'csv'
    else:
        content = _build_report_pdf_bytes(report_def, headers, rows, date_range_label)
        mime, ext = 'application/pdf', 'pdf'
        if content is None:
            messages.error(request, 'Could not generate the PDF for this report.')
            return redirect('reports_center')

    company = Company.objects.order_by('pk').first()
    subject = f"{report_def['label']} Report ({date_range_label}) - {company.name if company else 'NOA ERP'}"
    body = f"Please find attached the {report_def['label']} report for {date_range_label}.\n\nRegards,\n{company.name if company else 'NOA ERP'}"
    email = EmailMessage(subject, body, from_email=_from_email_for(request), to=[to_email])
    email.attach(_report_filename(report_def, ext), content, mime)
    try:
        email.send(fail_silently=False)
        messages.success(request, f"{report_def['label']} report emailed to {to_email}.")
    except Exception as e:
        messages.error(request, f"Could not send email: {e}")
    return redirect('reports_center')


@login_required
def reports_whatsapp(request):
    if not request.user.is_staff:
        return HttpResponseForbidden("You do not have permission to share reports.")
    from urllib.parse import quote, urlencode

    report_key, report_def, headers, rows, date_range_label = _get_report_data(request.GET)
    if not report_def:
        messages.error(request, 'Please choose a report first.')
        return redirect('reports_center')

    fmt = request.GET.get('format', 'pdf')
    download_params = {'report': report_key, 'format': fmt, 'preset': request.GET.get('preset', 'full')}
    if request.GET.get('preset') == 'custom':
        download_params['date_from'] = request.GET.get('date_from', '')
        download_params['date_to'] = request.GET.get('date_to', '')
    download_url = request.build_absolute_uri(reverse('reports_download') + '?' + urlencode(download_params))

    message = f"{report_def['label']} Report ({date_range_label}). View/download: {download_url}"
    digits = ''.join(ch for ch in request.GET.get('whatsapp_number', '') if ch.isdigit())
    wa_base = f"https://wa.me/{digits}" if digits else "https://wa.me/"
    return redirect(f"{wa_base}?text={quote(message)}")
